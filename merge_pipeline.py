"""
TikTok Live + Shop Data Merge Pipeline
"""

import os
import sys
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input_data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

FILES = {
    "live_product":   "live_product.csv",
    "shop_product":   "shop_product.csv",
    "live_perf":      "live_performance.csv",
    "shop_daily":     "shop_daily.csv",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_csv(key):
    path = os.path.join(INPUT_DIR, FILES[key])
    if not os.path.exists(path):
        print(f"  [WARNING] Missing file: {FILES[key]}. Please add it to the input_data/ folder.")
        return None
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip()
    return df


def clean_currency(series):
    """Strip $, commas, whitespace and coerce to float."""
    return (
        series.astype(str)
        .str.replace(r"[\$,\s]", "", regex=True)
        .pipe(pd.to_numeric, errors="coerce")
    )


def clean_numeric(series):
    """Remove commas/spaces and coerce to float."""
    return (
        series.astype(str)
        .str.replace(r"[,\s]", "", regex=True)
        .pipe(pd.to_numeric, errors="coerce")
    )


def parse_dates(series):
    """Handle MM/DD/YYYY, YYYY-MM-DD, and YYYY-MM-DD HH:MM formats."""
    return pd.to_datetime(series, errors="coerce")


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_live_product():
    df = load_csv("live_product")
    if df is None:
        return None
    currency_cols = [c for c in df.columns if any(k in c.lower() for k in ["gmv", "revenue"])]
    for col in currency_cols:
        df[col] = clean_currency(df[col])
    numeric_cols = [c for c in df.columns if any(k in c.lower() for k in ["sold", "order", "customer", "impression"])]
    for col in numeric_cols:
        df[col] = clean_numeric(df[col])
    ctr_cols = [c for c in df.columns if "ctr" in c.lower()]
    for col in ctr_cols:
        df[col] = (
            df[col].astype(str)
            .str.replace(r"[%\s]", "", regex=True)
            .pipe(pd.to_numeric, errors="coerce")
        )
    return df


def load_shop_product():
    df = load_csv("shop_product")
    if df is None:
        return None
    currency_cols = [c for c in df.columns if "gmv" in c.lower()]
    for col in currency_cols:
        df[col] = clean_currency(df[col])
    numeric_cols = [c for c in df.columns if any(k in c.lower() for k in ["sold", "order", "impression"])]
    for col in numeric_cols:
        df[col] = clean_numeric(df[col])
    ctr_cols = [c for c in df.columns if "ctr" in c.lower()]
    for col in ctr_cols:
        df[col] = (
            df[col].astype(str)
            .str.replace(r"[%\s]", "", regex=True)
            .pipe(pd.to_numeric, errors="coerce")
        )
    return df


def load_live_performance():
    df = load_csv("live_perf")
    if df is None:
        return None
    df["Start time"] = parse_dates(df.get("Start time", pd.Series(dtype=str)))
    df["_date"] = df["Start time"].dt.normalize()
    currency_cols = [c for c in df.columns if "gmv" in c.lower()]
    for col in currency_cols:
        df[col] = clean_currency(df[col])
    numeric_cols = [c for c in df.columns if any(k in c.lower() for k in ["order", "view", "follower"])]
    for col in numeric_cols:
        df[col] = clean_numeric(df[col])
    return df


def load_shop_daily():
    df = load_csv("shop_daily")
    if df is None:
        return None
    df["Date"] = parse_dates(df.get("Date", pd.Series(dtype=str)))
    df["_date"] = df["Date"].dt.normalize()
    currency_cols = [c for c in df.columns if "gmv" in c.lower()]
    for col in currency_cols:
        df[col] = clean_currency(df[col])
    numeric_cols = [c for c in df.columns if any(k in c.lower() for k in ["order", "sold", "impression"])]
    for col in numeric_cols:
        df[col] = clean_numeric(df[col])
    return df


# ---------------------------------------------------------------------------
# Merge 1 — Product-level JOIN
# ---------------------------------------------------------------------------

def merge_products(live_prod, shop_prod):
    if live_prod is None or shop_prod is None:
        print("  [SKIP] Product Comparison sheet skipped — one or both product files missing.")
        return pd.DataFrame()

    # Normalise join key
    live_prod["_pid"] = live_prod["Product ID"].astype(str).str.strip()
    shop_prod["_pid"] = shop_prod["Product ID"].astype(str).str.strip()

    # Pick the most likely column names (case-insensitive fallback helpers)
    def pick(df, *candidates):
        cols = {c.lower(): c for c in df.columns}
        for cand in candidates:
            if cand.lower() in cols:
                return cols[cand.lower()]
        return None

    live_gmv_col   = pick(live_prod, "Attributed GMV", "GMV")
    live_items_col = pick(live_prod, "Attributed items sold", "Items sold")
    live_ctr_col   = pick(live_prod, "CTR")
    live_name_col  = pick(live_prod, "Product name", "Product Name")

    shop_gmv_col   = pick(shop_prod, "GMV")
    shop_live_col  = pick(shop_prod, "Seller LIVE GMV", "LIVE GMV")
    shop_card_col  = pick(shop_prod, "Seller product card GMV", "Product card GMV")
    shop_orders_col= pick(shop_prod, "Orders")
    shop_ctr_col   = pick(shop_prod, "CTR")
    shop_name_col  = pick(shop_prod, "Product Name", "Product name")

    merged = live_prod.merge(shop_prod, on="_pid", how="outer", suffixes=("_live", "_shop"))

    # Build output with safe column access
    def safe_col(df, col, fallback=None):
        if col and col in df.columns:
            return df[col]
        if fallback and fallback in df.columns:
            return df[fallback]
        return pd.Series([None] * len(df))

    out = pd.DataFrame()
    out["Product ID"] = merged["_pid"]

    # Product name: prefer live, fall back to shop
    live_name_actual  = live_name_col + "_live"  if live_name_col  and live_name_col + "_live"  in merged.columns else live_name_col
    shop_name_actual  = shop_name_col + "_shop"  if shop_name_col  and shop_name_col + "_shop"  in merged.columns else shop_name_col
    out["Product Name"] = (
        merged.get(live_name_actual, pd.Series([None]*len(merged))).fillna(
            merged.get(shop_name_actual, pd.Series([None]*len(merged)))
        )
    )

    def resolve(base, suffix):
        suffixed = base + suffix if base else None
        if suffixed and suffixed in merged.columns:
            return merged[suffixed]
        if base and base in merged.columns:
            return merged[base]
        return pd.Series([None] * len(merged))

    out["Live Attributed GMV"]    = resolve(live_gmv_col,   "_live")
    out["Live Items Sold"]         = resolve(live_items_col, "_live")
    out["Shop Total GMV"]          = resolve(shop_gmv_col,   "_shop")
    out["Shop LIVE GMV"]           = resolve(shop_live_col,  "_shop")
    out["Shop Product Card GMV"]   = resolve(shop_card_col,  "_shop")
    out["Shop Orders"]             = resolve(shop_orders_col,"_shop")
    out["Live CTR (%)"]            = resolve(live_ctr_col,   "_live")
    out["Shop CTR (%)"]            = resolve(shop_ctr_col,   "_shop")

    out["Live GMV % of Shop Total"] = (
        out["Live Attributed GMV"] / out["Shop Total GMV"].replace(0, pd.NA) * 100
    ).round(2)

    out = out.sort_values("Live Attributed GMV", ascending=False, na_position="last")
    out = out.reset_index(drop=True)
    return out


# ---------------------------------------------------------------------------
# Merge 2 — Daily date JOIN
# ---------------------------------------------------------------------------

def merge_daily(live_perf, shop_daily):
    if live_perf is None and shop_daily is None:
        print("  [SKIP] Daily Overview sheet skipped — both daily files missing.")
        return pd.DataFrame()

    def pick(df, *candidates):
        cols = {c.lower(): c for c in df.columns}
        for cand in candidates:
            if cand.lower() in cols:
                return cols[cand.lower()]
        return None

    if live_perf is not None and shop_daily is not None:
        # Aggregate live sessions by date (multiple sessions on same day)
        lp = live_perf.copy()
        lp_name_col    = pick(lp, "Livestream", "Stream name", "Name")
        lp_gmv_col     = pick(lp, "Attributed GMV", "GMV")
        lp_orders_col  = pick(lp, "Attributed orders", "Orders")
        lp_views_col   = pick(lp, "Views")

        agg_dict = {}
        if lp_name_col:
            agg_dict[lp_name_col] = lambda x: " | ".join(x.dropna().astype(str))
        if lp_gmv_col:
            agg_dict[lp_gmv_col] = "sum"
        if lp_orders_col:
            agg_dict[lp_orders_col] = "sum"
        if lp_views_col:
            agg_dict[lp_views_col] = "sum"

        lp_agg = lp.groupby("_date", as_index=False).agg(agg_dict) if agg_dict else lp[["_date"]].drop_duplicates()

        sd = shop_daily.copy()
        sd_gmv_col        = pick(sd, "GMV")
        sd_live_gmv_col   = pick(sd, "LIVE-attributed GMV", "LIVE GMV")
        sd_orders_col     = pick(sd, "Orders")

        merged = sd.merge(lp_agg, on="_date", how="left", suffixes=("_shop", "_live"))
    elif live_perf is not None:
        merged = live_perf.copy()
        merged["_date"] = merged["_date"]
        sd_gmv_col = sd_live_gmv_col = sd_orders_col = None
        lp_name_col = pick(merged, "Livestream", "Stream name", "Name")
        lp_gmv_col  = pick(merged, "Attributed GMV", "GMV")
        lp_orders_col = pick(merged, "Attributed orders", "Orders")
        lp_views_col  = pick(merged, "Views")
    else:
        merged = shop_daily.copy()
        lp_name_col = lp_gmv_col = lp_orders_col = lp_views_col = None

    def resolve(base, suffix):
        if not base:
            return pd.Series([None] * len(merged))
        suffixed = base + suffix
        if suffixed in merged.columns:
            return merged[suffixed]
        if base in merged.columns:
            return merged[base]
        return pd.Series([None] * len(merged))

    out = pd.DataFrame()
    out["Date"] = merged["_date"].dt.date

    name_base = lp_name_col if live_perf is not None else None
    out["Live Session Name"] = resolve(name_base, "_live") if live_perf is not None else None

    out["Live GMV"]              = resolve(lp_gmv_col if live_perf is not None else None,   "_live")
    out["Live Orders"]           = resolve(lp_orders_col if live_perf is not None else None,"_live")
    out["Live Views"]            = resolve(lp_views_col if live_perf is not None else None, "_live")
    out["Shop Total GMV"]        = resolve(sd_gmv_col if shop_daily is not None else None,      "_shop")
    out["Shop LIVE-Attributed GMV"] = resolve(sd_live_gmv_col if shop_daily is not None else None,"_shop")
    out["Shop Orders"]           = resolve(sd_orders_col if shop_daily is not None else None,"_shop")

    out["Live Day vs Non-Live Day"] = out["Live Session Name"].apply(
        lambda x: "Live Day" if pd.notna(x) and str(x).strip() not in ("", "nan") else "No Live"
    )

    out = out.sort_values("Date", ascending=True).reset_index(drop=True)
    return out


# ---------------------------------------------------------------------------
# Sheet 3 — Live Summary
# ---------------------------------------------------------------------------

def build_live_summary(live_perf):
    if live_perf is None:
        print("  [SKIP] Live Summary sheet skipped — live_performance.csv missing.")
        return pd.DataFrame()

    df = live_perf.copy()
    # Drop internal helper column
    df = df.drop(columns=["_date"], errors="ignore")
    df = df.sort_values("Start time", ascending=True, na_position="last")
    df = df.reset_index(drop=True)

    # Build totals row for numeric columns
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    totals = {col: df[col].sum() for col in numeric_cols}
    totals_row = {col: "" for col in df.columns}
    totals_row.update(totals)

    # Label the totals row in the first column
    first_col = df.columns[0]
    totals_row[first_col] = "TOTAL"

    df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
    return df


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")   # dark navy
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL   = PatternFill("solid", fgColor="E8EAF6")
TOTAL_FONT   = Font(bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")


def format_sheet(ws):
    # Bold + coloured header row
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto column width (max 50)
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Alternate row shading + style totals row if present
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_total = str(row[0].value).upper() == "TOTAL" if row[0].value else False
        for cell in row:
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Freeze header row
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("\n=== TikTok Data Merge Pipeline ===\n")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Loading files...")
    live_prod  = load_live_product()
    shop_prod  = load_shop_product()
    live_perf  = load_live_performance()
    shop_daily = load_shop_daily()

    print("\nBuilding sheets...")
    sheet_product = merge_products(live_prod, shop_prod)
    sheet_daily   = merge_daily(live_perf, shop_daily)
    sheet_summary = build_live_summary(live_perf)

    today = date.today().strftime("%Y-%m-%d")
    output_path = os.path.join(OUTPUT_DIR, f"live_report_{today}.xlsx")

    print(f"\nWriting Excel report → {output_path}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        def write_sheet(df, name):
            if df.empty:
                placeholder = pd.DataFrame([["No data — check that the required input files are present."]])
                placeholder.to_excel(writer, sheet_name=name, index=False, header=False)
            else:
                df.to_excel(writer, sheet_name=name, index=False)

        write_sheet(sheet_product, "Product Comparison")
        write_sheet(sheet_daily,   "Daily Overview")
        write_sheet(sheet_summary, "Live Summary")

    # Apply formatting
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        if ws.max_row > 1:
            format_sheet(ws)
    wb.save(output_path)

    print(f"\n✓ Report saved to output/live_report_{today}.xlsx")
    print("=================================\n")


if __name__ == "__main__":
    main()
